import streamlit as st
import pandas as pd
import re
import io
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="Procesador de Pedidos", page_icon="📦", layout="wide")

st.markdown("""
<style>
[data-testid="stSidebar"] { background: #0f172a; }
[data-testid="stSidebar"] * { color: #cbd5e1 !important; }
</style>""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────
SPAIN = {"España", "Spain", "ES"}
NAC_ORDER = ["T_MIR", "T_AMZ", "T_C4", "T_MM", "T_PRIV"]

COUNTRY_SHEETS = {
    "Francia": ["ES-FR", "FR-FR"], "France": ["ES-FR", "FR-FR"],
    "Italia":  ["ES-IT", "IT-IT"], "Italy":  ["ES-IT", "IT-IT"],
    "Alemania":["ES-DE", "DE-DE"], "Germany":["ES-DE", "DE-DE"],
    "Portugal":["PT"],
    "Bélgica": ["BE", "BE - NL"],  "Belgium":    ["BE", "BE - NL"],
    "Países Bajos": ["NL", "BE - NL"], "Netherlands": ["NL", "BE - NL"],
    "Polonia": ["PL", "PL-SE"],    "Poland": ["PL", "PL-SE"],
    "Suecia":  ["SE", "PL-SE"],    "Sweden": ["SE", "PL-SE"],
}

MARKETPLACE_NAC_SHEET = {
    "carrefour":  "T_C4",
    "amazon":     "T_AMZ",
    "mediamarkt": "T_MM",
    "privalia":   "T_PRIV",
}

# ── Tarifa helpers ────────────────────────────────────────────────────────────
def parse_price(val):
    if pd.isna(val) or val == "" or val == 0: return None
    try: return float(str(val).replace("€","").replace(",",".").strip())
    except: return None

def normalize_sku(sku):
    s = str(sku).strip()
    if re.match(r'^A\d', s): return s
    s = re.sub(r'^[Ss]0*', '', s)
    return s.lstrip("0") or s

def split_skus(v):
    raw = str(v).strip()
    return [s for s in raw.split() if s] if raw and raw != "nan" else []

def expand_multi_sku_rows(df):
    rows = []
    for _, row in df.iterrows():
        skus = split_skus(row["Sku"])
        n = len(skus)
        for sku in skus:
            r = row.copy()
            r["_sku_orig"] = row["Sku"]
            r["Sku"] = sku
            r["_sku_norm"] = normalize_sku(sku)
            r["_multi"] = n > 1
            rows.append(r)
    return pd.DataFrame(rows)

def hoja1_to_rentabilidad(df):
    """
    Convert Hoja1 (orders export) into Rentabilidad format so analyze_row can process it.
    Hoja1 columns: ID, Cliente, mail, Entrega(País), Telefono, Total(Precio),
                   Pago(Marketplace), SKU, Cantidad, ..., Combination(Id Marketplace)
    Hoja2/Rentabilidad columns: Pedido, Fecha, Marketplace, Id Marketplace, País, Sku, Cant, Pedido.1
    Also handles comma-separated multi-SKU in the SKU column.
    """
    rows = []
    for _, r in df.iterrows():
        sku_raw = str(r.get("SKU", r.get("Sku", ""))).strip()
        # Split by comma or space for multi-SKU
        skus = [s.strip() for s in re.split(r"[, ]+", sku_raw) if s.strip() and s.strip() != "nan"]
        if not skus:
            continue
        pais = str(r.get("Entrega", r.get("País", ""))).strip()
        mkt  = str(r.get("Pago", r.get("Marketplace", ""))).strip()
        # Parse price from Total column (format: "17,90 €")
        precio_raw = r.get("Total", r.get("Pedido.1", ""))
        precio_str = str(precio_raw).replace("€","").replace(" ","").replace(" ","").replace(",",".").strip()
        combination = str(r.get("Combination", r.get("Id Marketplace", ""))).strip()
        n = len(skus)
        for sku in skus:
            rows.append({
                "Pedido":         r.get("ID", r.get("Pedido", "")),
                "Fecha":          r.get("Fecha", ""),
                "Marketplace":    mkt,
                "Id Marketplace": combination,
                "País":           pais,
                "Sku":            sku,
                "Cant":           r.get("Cantidad", r.get("Cant", 1)),
                "Pedido.1":       precio_str,
                "_sku_orig":      sku_raw,
                "_sku_norm":      normalize_sku(sku),
                "_multi":         n > 1,
            })
    return pd.DataFrame(rows)

def build_sheet_lookup(df, sheet_name):
    lookup = {}
    ref_cols = [c for c in df.columns if str(c).strip() == "REFERENCIA"]
    if not ref_cols: return lookup
    for _, r in df.iterrows():
        key = str(r[ref_cols[0]]).strip()
        if key:
            try: lookup[key] = {"min": float(r["PVP MIN."]), "pub": float(r["PVP PUB."]), "sheet": sheet_name}
            except: pass
    return lookup

@st.cache_data(show_spinner=False)
def load_tarifas(nac_bytes, inter_bytes):
    nac_raw   = pd.read_excel(io.BytesIO(nac_bytes),   sheet_name=None)
    inter_raw = pd.read_excel(io.BytesIO(inter_bytes),  sheet_name=None)
    nac_lkp   = {sh: build_sheet_lookup(df, sh) for sh, df in nac_raw.items()   if sh in NAC_ORDER}
    inter_lkp = {sh: build_sheet_lookup(df, sh) for sh, df in inter_raw.items()}
    return nac_lkp, inter_lkp

def nac_order_for(mkt):
    mkt_l = str(mkt).lower()
    primary = next((sh for kw, sh in MARKETPLACE_NAC_SHEET.items() if kw in mkt_l), "T_MIR")
    return [primary] + [s for s in NAC_ORDER if s != primary]

def get_pvp(sku_norm, pais, mkt, nac_lkp, inter_lkp):
    if pais in SPAIN:
        for sh in nac_order_for(mkt):
            lkp = nac_lkp.get(sh, {})
            if sku_norm in lkp: return lkp[sku_norm]
        return None
    for sh in COUNTRY_SHEETS.get(pais, []):
        lkp = inter_lkp.get(sh, {})
        if sku_norm in lkp: return lkp[sku_norm]
    return None

def analyze_row(row, nac_lkp, inter_lkp):
    sku   = row["_sku_norm"]
    pais  = str(row.get("País","")).strip()
    mkt   = str(row.get("Marketplace",""))
    price = parse_price(row.get("Pedido.1"))
    pvp   = get_pvp(sku, pais, mkt, nac_lkp, inter_lkp)

    if not pvp:
        return "❌ NO EN TARIFA", None, None, None, None, "CANCELAR"
    pvp_min, pvp_pub, sh = pvp["min"], pvp["pub"], pvp["sheet"]
    if price is None:
        return "⚠️ SIN PRECIO", pvp_min, pvp_pub, None, None, sh
    d_min = round(price - pvp_min, 2)
    d_pub = round(price - pvp_pub, 2)
    if price < pvp_min:
        status = "🔴 BAJO MÍNIMO" if abs(d_min) <= 10 else "🔴 BAJO MÍNIMO (>10€)"
    elif price == pvp_min:
        status = "🟡 EN MÍNIMO"
    else:
        status = "✅ OK"
    return status, pvp_min, pvp_pub, d_min, d_pub, sh

# ── Hoja1 helpers ─────────────────────────────────────────────────────────────
def clean_hoja1(df):
    df = df.drop(columns=[df.columns[0]])
    mask = df.apply(lambda row: all(str(v).strip() in ("--","-","nan","") for v in row), axis=1)
    return df[~mask].reset_index(drop=True)

def check_dupes_hoja1(df):
    if len(df.columns) < 15: return pd.DataFrame()
    cc, co = df.columns[2], df.columns[14]
    valid = df[(df[co].astype(str) != "--") & df[co].notna() & (df[co].astype(str) != "nan")]
    counts = valid.groupby([cc, co]).size()
    dup_keys = set(counts[counts > 1].index)
    return valid[valid.apply(lambda r: (r[cc], r[co]) in dup_keys, axis=1)].copy()

# ── Miravia helpers ───────────────────────────────────────────────────────────
def extract_comb_id(val):
    v = str(val).strip()
    return v[-13:] if len(v) >= 13 and v not in ("nan","--","") else ""

def check_dupes_miravia(df):
    if len(df.columns) < 14: return pd.DataFrame(), {}
    col = df.columns[13]
    ids = df[col].apply(extract_comb_id)
    df2 = df.copy(); df2["_id"] = ids
    valid = df2[df2["_id"] != ""]
    counts = valid["_id"].value_counts()
    dupes = valid[valid["_id"].isin(counts[counts > 1].index)].copy()
    return dupes, dict(zip(df.index, ids))

def cross_cancelados(df_h1, df_can):
    if df_can is None or len(df_h1.columns) < 14: return pd.DataFrame()
    col_comb = df_h1.columns[13]
    col_b    = df_can.columns[1]
    canceled = set(df_can[col_b].astype(str).str.strip())
    rows = []
    for _, row in df_h1.iterrows():
        mid = extract_comb_id(row.get(col_comb,""))
        if mid and mid in canceled:
            es = df_can[df_can[col_b].astype(str).str.strip() == mid]
            rows.append({
                "ID Pedido":   row.get("ID",""),
                "Combination": row.get(col_comb,""),
                "ID Extraído": mid,
                "SKU":         row.get("SKU",""),
                "Cliente":     row.get("Cliente",""),
                "Total":       row.get("Total",""),
                "Estado ES":   es["Estado"].values[0] if not es.empty else "",
                "Motivo":      es.iloc[0].get("Motivo de devolución: No se ha entregado al comprador","") if not es.empty else "",
            })
    return pd.DataFrame(rows)

# ── Email ─────────────────────────────────────────────────────────────────────
def get_smtp():
    try:
        return (st.secrets["correo"]["servidor_smtp"],
                int(st.secrets["correo"]["puerto"]),
                st.secrets["correo"]["usuario"],
                st.secrets["correo"]["password"])
    except: return None

def load_remitentes(b):
    df = pd.read_excel(io.BytesIO(b))
    df.columns = [c.strip() for c in df.columns]
    return df

def get_email_for(remitentes_df, canal):
    if remitentes_df is None: return None, None
    canal_col = next((c for c in remitentes_df.columns if "canal" in c.lower()), None)
    email_col = next((c for c in remitentes_df.columns if "mail"  in c.lower()), None)
    nombre_col= next((c for c in remitentes_df.columns if "nombre" in c.lower()), None)
    if not canal_col or not email_col: return None, None
    cl = str(canal).lower()
    for _, r in remitentes_df.iterrows():
        if str(r[canal_col]).lower() in cl or cl in str(r[canal_col]).lower():
            return str(r[email_col]).strip(), str(r[nombre_col]).strip() if nombre_col else canal
    return None, None

def send_email(to, asunto, df_lineas, smtp_cfg):
    server, port, user, pwd = smtp_cfg
    msg = MIMEMultipart("mixed")
    msg["Subject"] = asunto
    msg["From"]    = user
    msg["To"]      = to

    rows_html = "".join(
        f'<tr style="background:{"#f8fafc" if i%2==0 else "#fff"}">'
        + "".join(f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0">{v}</td>'
                  for v in [r.get("Pedido",""), r.get("Marketplace",""),
                             r.get("Id Marketplace",""), r.get("SKU Original",r.get("SKU","")),
                             r.get("País",""), r.get("Estado","")])
        + "</tr>"
        for i, (_, r) in enumerate(df_lineas.iterrows())
    )
    html = f"""<html><body style="font-family:Arial;color:#1e293b;max-width:750px;margin:0 auto">
    <div style="background:#1B2A4A;padding:18px 24px;border-radius:10px 10px 0 0">
      <h2 style="color:#fff;margin:0;font-size:19px">📦 {asunto}</h2>
    </div>
    <div style="padding:18px 24px;border:1px solid #e2e8f0;border-top:none;border-radius:0 0 10px 10px">
      <table style="width:100%;border-collapse:collapse;font-size:13px">
        <thead><tr style="background:#1B2A4A;color:#fff">
          <th style="padding:8px 12px;text-align:left">Pedido</th>
          <th style="padding:8px 12px;text-align:left">Marketplace</th>
          <th style="padding:8px 12px;text-align:left">ID Marketplace</th>
          <th style="padding:8px 12px;text-align:left">SKU</th>
          <th style="padding:8px 12px;text-align:left">País</th>
          <th style="padding:8px 12px;text-align:left">Estado</th>
        </tr></thead>
        <tbody>{rows_html}</tbody>
      </table>
      <p style="font-size:11px;color:#94a3b8;margin-top:16px">Generado por Procesador de Pedidos Turaco</p>
    </div></body></html>"""

    msg.attach(MIMEText(html, "html", "utf-8"))
    xlsx = build_excel([("Pedidos", df_lineas, None)])
    part = MIMEBase("application","octet-stream")
    part.set_payload(xlsx)
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", 'attachment; filename="pedidos.xlsx"')
    msg.attach(part)

    with smtplib.SMTP_SSL(server, port) as s:
        s.login(user, pwd)
        s.sendmail(user, to, msg.as_string())

# ── Excel export ──────────────────────────────────────────────────────────────
def build_excel(sections):
    wb = Workbook()
    HDR_FILL = PatternFill("solid", start_color="1B2A4A")
    HDR_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    STATUS_COLOR = {
        "✅ OK":                 "27AE60",
        "🟡 EN MÍNIMO":          "F39C12",
        "🔴 BAJO MÍNIMO":        "E74C3C",
        "🔴 BAJO MÍNIMO (>10€)": "C0392B",
        "❌ NO EN TARIFA":        "C0392B",
        "⚠️ SIN PRECIO":         "F39C12",
        "CANCELAR":              "C0392B",
    }
    first = True
    for sheet_name, df, summary in sections:
        ws = wb.active if first else wb.create_sheet()
        ws.title = sheet_name[:31]; first = False
        start = 1
        if summary:
            ws.cell(1,1,summary).font = Font(name="Arial", bold=True, size=11,
                color="27AE60" if "Sin" in summary or "✅" in summary else "C0392B")
            start = 3
        if df is None or len(df) == 0: continue
        for ci, col in enumerate(df.columns, 1):
            c = ws.cell(start, ci, str(col))
            c.font = HDR_FONT; c.fill = HDR_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for ri, (_, row) in enumerate(df.iterrows(), 1):
            for ci, val in enumerate(row.values, 1):
                ws.cell(start+ri, ci, val).font = Font(name="Arial", size=10)
            if "Estado" in df.columns:
                sci = list(df.columns).index("Estado") + 1
                sv  = str(row.get("Estado",""))
                for k, clr in STATUS_COLOR.items():
                    if k in sv:
                        c = ws.cell(start+ri, sci)
                        c.fill = PatternFill("solid", start_color=clr)
                        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                        break
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 20
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── Color status helper ───────────────────────────────────────────────────────
def color_status(val):
    m = {
        "✅ OK":                   "background-color:#d1fae5;color:#065f46",
        "🟡 EN MÍNIMO":            "background-color:#fef3c7;color:#92400e",
        "🔴 BAJO MÍNIMO":          "background-color:#fee2e2;color:#991b1b",
        "🔴 BAJO MÍNIMO (>10€)":   "background-color:#fecaca;color:#7f1d1d",
        "❌ NO EN TARIFA":          "background-color:#fce7f3;color:#831843",
        "⚠️ SIN PRECIO":           "background-color:#fef3c7;color:#92400e",
    }
    # Prefix match so both 🔴 variants get styled
    for k, v in m.items():
        if val == k: return v
    return ""

# ── Email widget (cancelados por canal) ───────────────────────────────────────
def cancelados_widget(df_tarifa, remitentes_df, key_prefix):
    """Gestión de cancelados sin re-renders: session_state para toda la selección."""
    if df_tarifa.empty:
        return
    mask = (df_tarifa["Estado"].str.startswith("❌") |
            df_tarifa["Estado"].str.startswith("🔴"))
    df_c = df_tarifa[mask].copy().reset_index(drop=True)
    if df_c.empty:
        return

    smtp_cfg = get_smtp()

    # ── session_state keys ────────────────────────────────────────────────────
    sk_sent = f"_sent_{key_prefix}"
    if sk_sent not in st.session_state: st.session_state[sk_sent] = set()

    # Build pedido labels (with estado color indicator)
    ESTADO_ICON = {"❌": "❌", "🔴": "🔴", "🟡": "🟡"}
    def pedido_label(row):
        estado = str(row.get("Estado",""))
        icon = next((v for k,v in ESTADO_ICON.items() if estado.startswith(k)), "")
        d = row.get("Dif vs Mín (€)","")
        dif = f" ({d}€)" if d != "" and d is not None and str(d) not in ("","nan") else ""
        return (
            f"{icon} {row.get('Pedido','')} · "
            f"{row.get('Marketplace','')} · "
            f"{row.get('Id Marketplace','')} · "
            f"SKU {row.get('SKU Original', row.get('SKU',''))} · "
            f"{row.get('País','')}{dif}"
        )

    todas_opciones = [pedido_label(row) for _, row in df_c.iterrows()]

    pendientes_idx = [i for i in range(len(df_c)) if i not in st.session_state[sk_sent]]

    # Build remitente options — sorted alphabetically
    rem_opciones = ["— elige remitente —"]
    rem_map: dict = {}
    if remitentes_df is not None:
        ec = next((c for c in remitentes_df.columns if "mail"  in c.lower()), None)
        cc = next((c for c in remitentes_df.columns if "canal" in c.lower()), None)
        nc = next((c for c in remitentes_df.columns if "nombre" in c.lower()), None)
        if ec and cc:
            entries = sorted([
                (f"{r[cc]} — {r[ec]}", (str(r[ec]).strip(), str(r[nc]).strip() if nc else str(r[cc])))
                for _, r in remitentes_df.iterrows()
            ], key=lambda x: x[0].lower())
            for lbl, val in entries:
                rem_opciones.append(lbl)
                rem_map[lbl] = val

    st.markdown("---")
    st.markdown("### 📋 Gestión de cancelados")

    # ── 1. Tabla sombreada ────────────────────────────────────────────────────
    rows_html = ""
    for i, (_, row) in enumerate(df_c.iterrows()):
        sent   = i in st.session_state[sk_sent]
        estado = str(row.get("Estado",""))
        d_min  = row.get("Dif vs Mín (€)","")
        bg          = "#f0fdf4" if sent else ("#fff7ed" if "🔴" in estado else "#fff1f2")
        badge_bg    = "#15803d" if sent else ("#92400e" if "🔴" in estado else "#831843")
        badge_txt   = "✅ Enviado" if sent else ("REVISAR" if "🔴" in estado else "CANCELAR")
        estado_bg   = "#fecaca" if "❌" in estado else ("#fed7aa" if "🔴" in estado else "#fef9c3")
        estado_col  = "#7f1d1d" if "❌" in estado else ("#92400e" if "🔴" in estado else "#713f12")
        rows_html += (
            f'<tr style="background:{bg}">'
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0;font-weight:600">{row.get("Pedido","")}</td>'
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0">{row.get("Marketplace","")}</td>'
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0">'
            f'<span style="background:{estado_bg};color:{estado_col};padding:2px 8px;border-radius:8px;font-size:11px;font-weight:700">{estado}</span></td>'
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0">{row.get("Id Marketplace","")}</td>'
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0"><code>{row.get("SKU Original", row.get("SKU",""))}</code></td>'
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0">{row.get("País","")}</td>'
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0">{d_min}</td>'
            f'<td style="padding:7px 12px;border-bottom:1px solid #e2e8f0">'
            f'<span style="background:{badge_bg};color:#fff;padding:2px 10px;border-radius:10px;font-size:11px;font-weight:700">{badge_txt}</span></td></tr>'
        )
    st.markdown(
        f'<table style="width:100%;border-collapse:collapse;font-size:13px;margin-bottom:16px">'
        f'<thead><tr style="background:#1B2A4A;color:#fff">'
        f'<th style="padding:8px 12px;text-align:left">Pedido</th>'
        f'<th style="padding:8px 12px;text-align:left">Marketplace</th>'
        f'<th style="padding:8px 12px;text-align:left">Estado</th>'
        f'<th style="padding:8px 12px;text-align:left">ID Marketplace</th>'
        f'<th style="padding:8px 12px;text-align:left">SKU</th>'
        f'<th style="padding:8px 12px;text-align:left">País</th>'
        f'<th style="padding:8px 12px;text-align:left">Dif Mín (€)</th>'
        f'<th style="padding:8px 12px;text-align:left">Enviado</th>'
        f'</tr></thead><tbody>{rows_html}</tbody></table>',
        unsafe_allow_html=True)

    # ── 2. Envío — todo dentro de st.form ────────────────────────────────────
    if not smtp_cfg:
        st.caption("⚠️ SMTP no configurado en Secrets")
        return

    st.markdown("**📧 Enviar pedidos pendientes:**")

    # Show pending result from previous submit (persisted across re-runs)
    sk_msg = f"_msg_{key_prefix}"
    if sk_msg in st.session_state:
        msg_type, msg_text = st.session_state.pop(sk_msg)
        if msg_type == "ok":    st.success(msg_text)
        elif msg_type == "warn": st.warning(msg_text)
        else:                    st.error(msg_text)

    with st.form(key=f"{key_prefix}_form", clear_on_submit=False):
        rem_sel     = st.selectbox("Destinatario", rem_opciones)
        asunto      = st.text_input("Asunto", value="Cancelados")
        pedidos_sel = st.multiselect(
            "Pedidos a incluir",
            options=todas_opciones,
            default=todas_opciones,
        )
        submitted = st.form_submit_button("📤 Enviar email", type="primary")

    if submitted:
        email_dest, _ = rem_map.get(rem_sel, (None, None))
        if not email_dest:
            st.session_state[sk_msg] = ("warn", "⚠️ Elige un destinatario válido")
        elif not pedidos_sel:
            st.session_state[sk_msg] = ("warn", "⚠️ Selecciona al menos un pedido")
        else:
            idx_sel  = [i for i, lbl in enumerate(todas_opciones) if lbl in pedidos_sel]
            df_envio = df_c.iloc[idx_sel].copy()
            try:
                send_email(email_dest, asunto, df_envio, smtp_cfg)
                for i in idx_sel:
                    st.session_state[sk_sent].add(i)
                st.session_state[sk_msg] = ("ok", f"✅ Enviado a {email_dest} ({len(df_envio)} pedidos)")
            except Exception as e:
                st.session_state[sk_msg] = ("err", f"❌ {e}")
        st.rerun()

    if st.session_state[sk_sent]:
        if st.button("↩ Limpiar enviados", key=f"{key_prefix}_clear"):
            st.session_state[sk_sent] = set()

# ═══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## 📦 Procesador Pedidos")
    st.markdown("---")
    proceso = st.radio("**Proceso**",
        ["🛒  Pago aceptado", "🏪  Pago aceptado Miravia"], key="proceso_radio")
    es_miravia = "Miravia" in proceso

    st.markdown("---")
    st.markdown("### Tarifas (obligatorio)")
    nac_file   = st.file_uploader("Tarifa Nacional (.xlsx)",       type="xlsx", key="nac")
    inter_file = st.file_uploader("Tarifa Internacional (.xlsx)",  type="xlsx", key="inter")

    st.markdown("---")
    st.markdown("### 📧 Remitentes")
    remitentes_file = st.file_uploader(
        "Fichero remitentes (.xlsx) — columnas: Canal, Email, Nombre",
        type="xlsx", key="rem")

    st.markdown("---")
    if es_miravia:
        st.markdown("### Ficheros Miravia")
        miravia_file    = st.file_uploader("PagoAceptadoMiravia.xlsx", type="xlsx", key="miravia")
        cancelados_file = st.file_uploader("Fichero ES... cancelados (.xlsx)", type="xlsx", key="cancelados")
        libro_file = None
    else:
        st.markdown("### Fichero de pedidos")
        libro_file = st.file_uploader(
            "Rentabilidad (.xlsx) — Turaco / Jabiru", type="xlsx", key="libro")
        miravia_file = cancelados_file = None

    st.markdown("---")
    run_btn = st.button("▶ Procesar", type="primary", use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
st.title("📦 Procesador de Pedidos")

# Load remitentes
remitentes_df = None
if remitentes_file:
    try:
        remitentes_df = load_remitentes(remitentes_file.read())
        st.sidebar.success(f"✅ Remitentes: {len(remitentes_df)} canales")
    except Exception as e:
        st.sidebar.warning(f"⚠️ Remitentes: {e}")

# Persist run state so form submits (re-runs) don't lose context
if run_btn:
    st.session_state["_ran"] = True

if not run_btn and not st.session_state.get("_ran"):
    smtp_ok = get_smtp() is not None
    st.info("👈 Selecciona el proceso, sube los ficheros y pulsa **Procesar**.")
    st.markdown(f"""
**Proceso A — Pago aceptado:** análisis tarifa + duplicados Hoja1 + gestión cancelados  
**Proceso B — Pago aceptado Miravia:** cruce cancelados (DERECHA 13 de Combination) + duplicados  
**SMTP:** {'✅ configurado en Secrets' if smtp_ok else '⚠️ no configurado — añade `[correo]` en Secrets para enviar emails'}  
**Remitentes:** {'✅ cargado' if remitentes_df is not None else '⚠️ no cargado — sube el fichero para ver emails por canal'}
    """)
    st.stop()

# Validate
errors = []
if not nac_file:   errors.append("Tarifa Nacional")
if not inter_file: errors.append("Tarifa Internacional")
if not es_miravia and not libro_file:   errors.append("Fichero Rentabilidad")
if es_miravia     and not miravia_file: errors.append("PagoAceptadoMiravia.xlsx")
if errors:
    st.error(f"Faltan ficheros: {', '.join(errors)}"); st.stop()

# Load tarifas
with st.spinner("Cargando tarifas..."):
    nac_lkp, inter_lkp = load_tarifas(nac_file.read(), inter_file.read())
n_nac   = sum(len(v) for v in nac_lkp.values())
n_inter = sum(len(v) for v in inter_lkp.values())
st.success(f"✅ Tarifas — Nacional: {n_nac:,} SKUs | Internacional: {n_inter:,} refs")

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESO A
# ═══════════════════════════════════════════════════════════════════════════════
if not es_miravia:
    with st.spinner("Procesando..."):
        sheets = pd.read_excel(io.BytesIO(libro_file.read()), sheet_name=None)
        h1_raw = sheets.get("Hoja1", pd.DataFrame())
        h2_raw = sheets.get("Hoja2", pd.DataFrame())
        if "Hoja2" not in sheets and len(sheets) > 1:
            h2_raw = list(sheets.values())[1]

        h1_clean  = clean_hoja1(h1_raw) if not h1_raw.empty else pd.DataFrame()
        dupes_h1  = check_dupes_hoja1(h1_clean) if not h1_clean.empty else pd.DataFrame()

        # If no Hoja2, derive rentabilidad data from Hoja1 directly
        fuente_tarifa = "Hoja2"
        if h2_raw.empty and not h1_clean.empty:
            h2_raw = hoja1_to_rentabilidad(h1_clean)
            fuente_tarifa = "Hoja1"

        df_tarifa   = pd.DataFrame()
        multi_count = 0
        if not h2_raw.empty:
            # If already expanded by hoja1_to_rentabilidad, skip expand; else expand
            if fuente_tarifa == "Hoja1":
                df_exp = h2_raw  # already expanded and normalized
            else:
                df_exp = expand_multi_sku_rows(h2_raw)
            multi_count = int(df_exp["_multi"].sum()) if "_multi" in df_exp.columns else 0
            results = []
            for _, row in df_exp.iterrows():
                status, pvp_min, pvp_pub, d_min, d_pub, sh = analyze_row(row, nac_lkp, inter_lkp)
                results.append({
                    "Pedido":          row.get("Pedido",""),
                    "Fecha":           row.get("Fecha",""),
                    "Marketplace":     row.get("Marketplace",""),
                    "Id Marketplace":  row.get("Id Marketplace",""),
                    "País":            row.get("País",""),
                    "SKU Original":    row.get("_sku_orig", row.get("Sku","")),
                    "SKU":             row.get("Sku",""),
                    "SKU Norm.":       row.get("_sku_norm",""),
                    "Multi-SKU":       "✔" if row.get("_multi") else "",
                    "Cant":            row.get("Cant",""),
                    "Precio Pedido (€)": parse_price(row.get("Pedido.1")),
                    "Hoja Tarifa":     sh,
                    "PVP Mín (€)":     pvp_min,
                    "PVP Pub (€)":     pvp_pub,
                    "Dif vs Mín (€)":  d_min,
                    "Dif vs Pub (€)":  d_pub,
                    "Estado":          status,
                })
            df_tarifa = pd.DataFrame(results)

    st.markdown("---")
    st.markdown("## 🛒 Pago aceptado — Resultados")

    ok     = (df_tarifa["Estado"] == "✅ OK").sum() if not df_tarifa.empty else 0
    warn   = (df_tarifa["Estado"].str.startswith("🔴").sum() + df_tarifa["Estado"].str.startswith("🟡").sum()) if not df_tarifa.empty else 0
    cancel = df_tarifa["Estado"].str.startswith("❌").sum() if not df_tarifa.empty else 0

    c1,c2,c3,c4,c5 = st.columns(5)
    c1.metric("📋 Hoja1 filas",  len(h1_clean))
    c2.metric("⚠️ Duplicados",   len(dupes_h1))
    c3.metric("✅ OK tarifa",     ok)
    c4.metric("🔴 Bajo mínimo",  warn)
    c5.metric("❌ No en tarifa", cancel)

    tab1, tab2, tab3 = st.tabs(["💰 Análisis Tarifa", "🔍 Duplicados Hoja1", "📄 Hoja1 limpia"])
    with tab1:
        if not df_tarifa.empty:
            if fuente_tarifa == "Hoja1":
                st.info("ℹ️ Análisis realizado sobre Hoja1 (no existe Hoja2 en el fichero)")
            if multi_count > 0:
                st.info(f"🔀 {multi_count} líneas generadas por expansión de pedidos multi-SKU")
            st.dataframe(df_tarifa.style.map(color_status, subset=["Estado"]),
                         use_container_width=True, hide_index=True)
        else:
            st.info("No hay datos para análisis de tarifa.")
    with tab2:
        if len(dupes_h1) > 0:
            st.warning(f"⚠️ {len(dupes_h1)} filas duplicadas (col C mail + col O Marketplace Order ID)")
            st.dataframe(dupes_h1, use_container_width=True, hide_index=True)
        else:
            st.success("✅ Sin duplicados en Hoja1")
    with tab3:
        if not h1_clean.empty:
            st.dataframe(h1_clean, use_container_width=True, hide_index=True)

    # Cancelados widget (always visible, outside tabs)
    cancelados_widget(df_tarifa, remitentes_df, "procA")

    st.markdown("---")
    excel = build_excel([
        ("Análisis Tarifa", df_tarifa, None),
        ("Duplicados Hoja1", dupes_h1 if len(dupes_h1) > 0 else None,
         "Sin duplicados" if len(dupes_h1) == 0 else None),
        ("Hoja1 limpia", h1_clean, None),
    ])
    st.download_button("⬇️ Descargar Excel completo", data=excel,
        file_name=f"PagoAceptado_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")

# ═══════════════════════════════════════════════════════════════════════════════
# PROCESO B — Miravia
# ═══════════════════════════════════════════════════════════════════════════════
else:
    with st.spinner("Procesando Miravia..."):
        mir_sheets = pd.read_excel(io.BytesIO(miravia_file.read()), sheet_name=None)
        h1_raw = mir_sheets.get("Hoja1", pd.DataFrame())
        h2_raw = mir_sheets.get("Hoja2", pd.DataFrame())

        df_cancelados = None
        if cancelados_file:
            can_sheets = pd.read_excel(io.BytesIO(cancelados_file.read()), sheet_name=None)
            df_cancelados = list(can_sheets.values())[0]

        h1_clean = clean_hoja1(h1_raw) if not h1_raw.empty else pd.DataFrame()
        dupes_comb, id_map = check_dupes_miravia(h1_clean) if not h1_clean.empty else (pd.DataFrame(), {})
        df_cancel_match = cross_cancelados(h1_clean, df_cancelados)

        df_tarifa   = pd.DataFrame()
        multi_count = 0
        if not h2_raw.empty:
            df_exp = expand_multi_sku_rows(h2_raw)
            multi_count = int(df_exp["_multi"].sum())
            results = []
            for _, row in df_exp.iterrows():
                status, pvp_min, pvp_pub, d_min, d_pub, sh = analyze_row(row, nac_lkp, inter_lkp)
                results.append({
                    "Pedido": row.get("Pedido",""), "SKU": row.get("Sku",""),
                    "SKU Norm.": row.get("_sku_norm",""), "Multi-SKU": "✔" if row.get("_multi") else "",
                    "País": row.get("País",""), "Marketplace": row.get("Marketplace",""),
                    "Precio Pedido (€)": parse_price(row.get("Pedido.1")),
                    "Hoja Tarifa": sh, "PVP Mín (€)": pvp_min, "PVP Pub (€)": pvp_pub,
                    "Dif vs Mín (€)": d_min, "Estado": status,
                })
            df_tarifa = pd.DataFrame(results)

    st.markdown("---")
    st.markdown("## 🏪 Pago aceptado Miravia — Resultados")

    c1,c2,c3,c4 = st.columns(4)
    c1.metric("📋 Pedidos Miravia",     len(h1_clean))
    c2.metric("🔴 Cancelados match",    len(df_cancel_match))
    c3.metric("⚠️ Dupl. Combination",  len(dupes_comb))
    c4.metric("💰 Tarifa analizados",   len(df_tarifa))

    tab1, tab2, tab3, tab4 = st.tabs(["🔴 Cancelados ES", "⚠️ Duplicados Combination",
                                       "💰 Análisis Tarifa", "📄 Hoja1 limpia"])
    with tab1:
        if cancelados_file is None:
            st.info("No se ha subido el fichero ES de cancelados.")
        elif len(df_cancel_match) > 0:
            st.error(f"🔴 {len(df_cancel_match)} pedidos Miravia en el fichero de cancelados")
            st.dataframe(df_cancel_match, use_container_width=True, hide_index=True)
        else:
            n_ids = sum(1 for v in id_map.values() if v)
            n_can = len(df_cancelados) if df_cancelados is not None else 0
            st.success(f"✅ Sin cancelados ({n_ids} IDs cruzados contra {n_can:,})")
    with tab2:
        if len(dupes_comb) > 0:
            st.warning(f"⚠️ {len(dupes_comb)} filas con ID Combination duplicado")
            st.dataframe(dupes_comb, use_container_width=True, hide_index=True)
        else:
            st.success("✅ Sin IDs duplicados en columna Combination")
    with tab3:
        if not df_tarifa.empty:
            if multi_count > 0:
                st.info(f"🔀 {multi_count} líneas por expansión multi-SKU")
            st.dataframe(df_tarifa.style.map(color_status, subset=["Estado"]),
                         use_container_width=True, hide_index=True)
        else:
            st.info("Sin datos en Hoja2 (habitual en Miravia).")
    with tab4:
        st.dataframe(h1_clean, use_container_width=True, hide_index=True)

    # Cancelados widget (always visible, outside tabs)
    cancelados_widget(df_tarifa, remitentes_df, "procB")

    st.markdown("---")
    excel = build_excel([
        ("Cancelados Match",      df_cancel_match if len(df_cancel_match) > 0 else None,
         "Sin cancelados" if len(df_cancel_match) == 0 else None),
        ("Duplicados Combination", dupes_comb if len(dupes_comb) > 0 else None,
         "Sin duplicados" if len(dupes_comb) == 0 else None),
        ("Análisis Tarifa",       df_tarifa if not df_tarifa.empty else None, None),
        ("Hoja1 limpia",          h1_clean, None),
    ])
    st.download_button("⬇️ Descargar Excel completo", data=excel,
        file_name=f"PagoAceptadoMiravia_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", type="primary")
